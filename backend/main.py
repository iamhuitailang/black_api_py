import os
import json
import io
import uuid
from datetime import datetime, timedelta
from typing import Optional, List
from math import radians, sin, cos, sqrt, atan2

from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File, Form, Query
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm

from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime, Boolean, ForeignKey, Text
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, Session, relationship

from pydantic import BaseModel, EmailStr
from passlib.context import CryptContext
from jose import JWTError, jwt

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

SECRET_KEY = "pet-finder-secret-key-2026"
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE_URL = f"sqlite:///{os.path.join(BASE_DIR, 'pet_finder.db')}"
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/api/login")


class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String(50), unique=True, index=True, nullable=False)
    email = Column(String(100), unique=True, index=True, nullable=False)
    phone = Column(String(20))
    address = Column(String(255))
    hashed_password = Column(String(255), nullable=False)
    is_admin = Column(Boolean, default=False)
    created_at = Column(DateTime, default=datetime.utcnow)
    pets = relationship("Pet", back_populates="owner")


class Pet(Base):
    __tablename__ = "pets"
    id = Column(Integer, primary_key=True, index=True)
    owner_id = Column(Integer, ForeignKey("users.id"), nullable=False)
    name = Column(String(100), nullable=False)
    species = Column(String(50), nullable=False)
    breed = Column(String(100), nullable=False)
    color = Column(String(100), nullable=False)
    photo = Column(String(255))
    chip_number = Column(String(100))
    is_neutered = Column(Boolean, default=False)
    contact_phone = Column(String(20), nullable=False)
    description = Column(Text)
    created_at = Column(DateTime, default=datetime.utcnow)
    owner = relationship("User", back_populates="pets")
    lost_records = relationship("LostRecord", back_populates="pet")


class LostRecord(Base):
    __tablename__ = "lost_records"
    id = Column(Integer, primary_key=True, index=True)
    pet_id = Column(Integer, ForeignKey("pets.id"), nullable=False)
    lost_time = Column(DateTime, nullable=False)
    lost_location = Column(String(255), nullable=False)
    lat = Column(Float)
    lng = Column(Float)
    features = Column(Text, nullable=False)
    reward = Column(Float, default=0)
    status = Column(String(20), default="lost")
    found_time = Column(DateTime)
    found_location = Column(String(255))
    created_at = Column(DateTime, default=datetime.utcnow)
    pet = relationship("Pet", back_populates="lost_records")
    matches = relationship("FoundMatch", back_populates="lost_record")
    timeline = relationship("ReunionTimeline", back_populates="lost_record")


class FoundMatch(Base):
    __tablename__ = "found_matches"
    id = Column(Integer, primary_key=True, index=True)
    lost_record_id = Column(Integer, ForeignKey("lost_records.id"), nullable=False)
    finder_name = Column(String(100), nullable=False)
    finder_phone = Column(String(20), nullable=False)
    species = Column(String(50), nullable=False)
    breed = Column(String(100), nullable=False)
    color = Column(String(100), nullable=False)
    found_location = Column(String(255), nullable=False)
    lat = Column(Float)
    lng = Column(Float)
    photo = Column(String(255))
    features = Column(Text)
    match_score = Column(Float, default=0)
    created_at = Column(DateTime, default=datetime.utcnow)
    lost_record = relationship("LostRecord", back_populates="matches")


class ReunionTimeline(Base):
    __tablename__ = "reunion_timeline"
    id = Column(Integer, primary_key=True, index=True)
    lost_record_id = Column(Integer, ForeignKey("lost_records.id"), nullable=False)
    event_type = Column(String(50), nullable=False)
    description = Column(Text, nullable=False)
    event_time = Column(DateTime, default=datetime.utcnow)
    photo = Column(String(255))
    lost_record = relationship("LostRecord", back_populates="timeline")


Base.metadata.create_all(bind=engine)


class UserCreate(BaseModel):
    username: str
    email: EmailStr
    password: str
    phone: Optional[str] = None
    address: Optional[str] = None


class UserResponse(BaseModel):
    id: int
    username: str
    email: str
    phone: Optional[str] = None
    address: Optional[str] = None
    is_admin: bool

    class Config:
        from_attributes = True


class Token(BaseModel):
    access_token: str
    token_type: str
    user: UserResponse


class PetCreate(BaseModel):
    name: str
    species: str
    breed: str
    color: str
    chip_number: Optional[str] = None
    is_neutered: bool = False
    contact_phone: str
    description: Optional[str] = None


class PetResponse(BaseModel):
    id: int
    owner_id: int
    name: str
    species: str
    breed: str
    color: str
    photo: Optional[str] = None
    chip_number: Optional[str] = None
    is_neutered: bool
    contact_phone: str
    description: Optional[str] = None
    created_at: datetime

    class Config:
        from_attributes = True


class LostRecordCreate(BaseModel):
    pet_id: int
    lost_time: str
    lost_location: str
    lat: Optional[float] = None
    lng: Optional[float] = None
    features: str
    reward: float = 0


class LostRecordResponse(BaseModel):
    id: int
    pet_id: int
    pet_name: str
    pet_species: str
    pet_breed: str
    pet_color: str
    pet_photo: Optional[str] = None
    owner_name: str
    owner_phone: str
    lost_time: datetime
    lost_location: str
    lat: Optional[float] = None
    lng: Optional[float] = None
    features: str
    reward: float
    status: str
    found_time: Optional[datetime] = None
    found_location: Optional[str] = None
    created_at: datetime

    class Config:
        from_attributes = True


class FoundMatchCreate(BaseModel):
    finder_name: str
    finder_phone: str
    species: str
    breed: str
    color: str
    found_location: str
    lat: Optional[float] = None
    lng: Optional[float] = None
    features: Optional[str] = None


class FoundMatchResponse(BaseModel):
    id: int
    lost_record_id: int
    finder_name: str
    finder_phone: str
    species: str
    breed: str
    color: str
    found_location: str
    lat: Optional[float] = None
    lng: Optional[float] = None
    photo: Optional[str] = None
    features: Optional[str] = None
    match_score: float
    created_at: datetime
    matched_lost: Optional[LostRecordResponse] = None

    class Config:
        from_attributes = True


class TimelineEvent(BaseModel):
    event_type: str
    description: str
    event_time: datetime
    photo: Optional[str] = None


class LostRecordDetail(LostRecordResponse):
    timeline: List[TimelineEvent] = []
    matches: List[FoundMatchResponse] = []


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)


def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)


def create_access_token(data: dict, expires_delta: Optional[timedelta] = None) -> str:
    to_encode = data.copy()
    expire = datetime.utcnow() + (expires_delta or timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES))
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)


def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)) -> User:
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="无法验证凭证",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: int = int(payload.get("sub"))
        if user_id is None:
            raise credentials_exception
    except (JWTError, ValueError):
        raise credentials_exception
    user = db.query(User).filter(User.id == user_id).first()
    if user is None:
        raise credentials_exception
    return user


def get_admin_user(current_user: User = Depends(get_current_user)) -> User:
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="需要管理员权限")
    return current_user


def haversine_distance(lat1, lng1, lat2, lng2):
    R = 6371.0
    if None in [lat1, lng1, lat2, lng2]:
        return None
    lat1, lng1, lat2, lng2 = map(radians, [lat1, lng1, lat2, lng2])
    dlat = lat2 - lat1
    dlng = lng2 - lng1
    a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlng / 2) ** 2
    c = 2 * atan2(sqrt(a), sqrt(1 - a))
    return R * c


app = FastAPI(title="小区宠物登记与寻宠平台")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")
app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "..", "frontend", "static")), name="static")


def save_upload_file(file: UploadFile) -> Optional[str]:
    if not file or not file.filename:
        return None
    ext = os.path.splitext(file.filename)[1]
    filename = f"{uuid.uuid4().hex}{ext}"
    filepath = os.path.join(UPLOAD_DIR, filename)
    with open(filepath, "wb") as f:
        content = file.file.read()
        f.write(content)
    return f"/uploads/{filename}"


@app.get("/")
def read_root():
    return FileResponse(os.path.join(BASE_DIR, "..", "frontend", "index.html"))


@app.post("/api/register", response_model=Token)
def register(user_data: UserCreate, db: Session = Depends(get_db)):
    existing = db.query(User).filter((User.username == user_data.username) | (User.email == user_data.email)).first()
    if existing:
        raise HTTPException(status_code=400, detail="用户名或邮箱已存在")
    user = User(
        username=user_data.username,
        email=user_data.email,
        phone=user_data.phone,
        address=user_data.address,
        hashed_password=get_password_hash(user_data.password),
        is_admin=False,
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    if db.query(User).count() == 1:
        user.is_admin = True
        db.commit()
        db.refresh(user)
    access_token = create_access_token(data={"sub": str(user.id)})
    return Token(access_token=access_token, token_type="bearer", user=UserResponse.model_validate(user))


@app.post("/api/login", response_model=Token)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == form_data.username).first()
    if not user or not verify_password(form_data.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    access_token = create_access_token(data={"sub": str(user.id)})
    return Token(access_token=access_token, token_type="bearer", user=UserResponse.model_validate(user))


@app.get("/api/me", response_model=UserResponse)
def get_me(current_user: User = Depends(get_current_user)):
    return UserResponse.model_validate(current_user)


@app.post("/api/pets", response_model=PetResponse)
async def create_pet(
    name: str = Form(...),
    species: str = Form(...),
    breed: str = Form(...),
    color: str = Form(...),
    chip_number: Optional[str] = Form(None),
    is_neutered: bool = Form(False),
    contact_phone: str = Form(...),
    description: Optional[str] = Form(None),
    photo: Optional[UploadFile] = File(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    photo_path = save_upload_file(photo) if photo else None
    pet = Pet(
        owner_id=current_user.id,
        name=name,
        species=species,
        breed=breed,
        color=color,
        photo=photo_path,
        chip_number=chip_number,
        is_neutered=is_neutered,
        contact_phone=contact_phone,
        description=description,
    )
    db.add(pet)
    db.commit()
    db.refresh(pet)
    return PetResponse.model_validate(pet)


@app.get("/api/pets", response_model=List[PetResponse])
def list_my_pets(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    pets = db.query(Pet).filter(Pet.owner_id == current_user.id).all()
    return [PetResponse.model_validate(p) for p in pets]


@app.get("/api/pets/all", response_model=List[PetResponse])
def list_all_pets(db: Session = Depends(get_db), current_user: User = Depends(get_admin_user)):
    pets = db.query(Pet).all()
    return [PetResponse.model_validate(p) for p in pets]


def build_lost_response(record: LostRecord) -> LostRecordResponse:
    pet = record.pet
    owner = pet.owner
    return LostRecordResponse(
        id=record.id,
        pet_id=record.pet_id,
        pet_name=pet.name,
        pet_species=pet.species,
        pet_breed=pet.breed,
        pet_color=pet.color,
        pet_photo=pet.photo,
        owner_name=owner.username,
        owner_phone=pet.contact_phone,
        lost_time=record.lost_time,
        lost_location=record.lost_location,
        lat=record.lat,
        lng=record.lng,
        features=record.features,
        reward=record.reward,
        status=record.status,
        found_time=record.found_time,
        found_location=record.found_location,
        created_at=record.created_at,
    )


@app.post("/api/lost", response_model=LostRecordResponse)
def create_lost_record(data: LostRecordCreate, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    pet = db.query(Pet).filter(Pet.id == data.pet_id, Pet.owner_id == current_user.id).first()
    if not pet:
        raise HTTPException(status_code=404, detail="宠物不存在或无权操作")
    try:
        lost_time = datetime.fromisoformat(data.lost_time.replace("Z", "+00:00"))
    except Exception:
        lost_time = datetime.utcnow()
    record = LostRecord(
        pet_id=data.pet_id,
        lost_time=lost_time,
        lost_location=data.lost_location,
        lat=data.lat,
        lng=data.lng,
        features=data.features,
        reward=data.reward,
        status="lost",
    )
    db.add(record)
    db.commit()
    db.refresh(record)

    db.add(ReunionTimeline(
        lost_record_id=record.id,
        event_type="走失发布",
        description=f"{current_user.username} 发布了 {pet.name} 的走失启事，走失地点：{data.lost_location}",
        event_time=datetime.utcnow(),
    ))
    db.commit()

    return build_lost_response(record)


@app.get("/api/lost", response_model=List[LostRecordResponse])
def list_lost_records(status: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(LostRecord).order_by(LostRecord.created_at.desc())
    if status:
        query = query.filter(LostRecord.status == status)
    records = query.all()
    return [build_lost_response(r) for r in records]


@app.get("/api/lost/carousel", response_model=List[LostRecordResponse])
def carousel_lost_records(db: Session = Depends(get_db)):
    records = db.query(LostRecord).filter(LostRecord.status == "lost").order_by(LostRecord.created_at.desc()).limit(10).all()
    return [build_lost_response(r) for r in records]


@app.get("/api/lost/{record_id}", response_model=LostRecordDetail)
def get_lost_detail(record_id: int, db: Session = Depends(get_db)):
    record = db.query(LostRecord).filter(LostRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    resp = build_lost_response(record)
    detail = LostRecordDetail(**resp.model_dump())
    detail.timeline = [TimelineEvent(
        event_type=t.event_type,
        description=t.description,
        event_time=t.event_time,
        photo=t.photo,
    ) for t in sorted(record.timeline, key=lambda x: x.event_time)]
    detail.matches = [FoundMatchResponse(
        id=m.id,
        lost_record_id=m.lost_record_id,
        finder_name=m.finder_name,
        finder_phone=m.finder_phone,
        species=m.species,
        breed=m.breed,
        color=m.color,
        found_location=m.found_location,
        lat=m.lat,
        lng=m.lng,
        photo=m.photo,
        features=m.features,
        match_score=m.match_score,
        created_at=m.created_at,
    ) for m in record.matches]
    return detail


@app.post("/api/found", response_model=List[FoundMatchResponse])
async def create_found_match(
    finder_name: str = Form(...),
    finder_phone: str = Form(...),
    species: str = Form(...),
    breed: str = Form(...),
    color: str = Form(...),
    found_location: str = Form(...),
    lat: Optional[float] = Form(None),
    lng: Optional[float] = Form(None),
    features: Optional[str] = Form(None),
    photo: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
):
    photo_path = save_upload_file(photo) if photo else None

    lost_records = db.query(LostRecord).filter(LostRecord.status == "lost").all()

    matched = []
    for lr in lost_records:
        pet = lr.pet
        score = 0.0

        if pet.species.lower() == species.lower():
            score += 35
        if pet.breed.lower() == breed.lower():
            score += 25
        elif any(b in pet.breed.lower() for b in [breed.lower()]) or any(b in breed.lower() for b in [pet.breed.lower()]):
            score += 10

        if pet.color.lower() == color.lower():
            score += 25
        elif any(c in pet.color.lower() for c in [color.lower()]) or any(c in color.lower() for c in [pet.color.lower()]):
            score += 10

        dist = haversine_distance(lat, lng, lr.lat, lr.lng)
        if dist is not None and dist <= 3:
            score += 15 * (1 - dist / 3)
        elif lr.lost_location and found_location:
            if lr.lost_location[:3] == found_location[:3]:
                score += 10

        if score >= 35:
            fm = FoundMatch(
                lost_record_id=lr.id,
                finder_name=finder_name,
                finder_phone=finder_phone,
                species=species,
                breed=breed,
                color=color,
                found_location=found_location,
                lat=lat,
                lng=lng,
                photo=photo_path,
                features=features,
                match_score=round(score, 1),
            )
            db.add(fm)
            matched.append((fm, lr, score))

    db.commit()

    result = []
    for fm, lr, score in sorted(matched, key=lambda x: -x[2]):
        db.refresh(fm)
        resp = FoundMatchResponse(
            id=fm.id,
            lost_record_id=fm.lost_record_id,
            finder_name=fm.finder_name,
            finder_phone=fm.finder_phone,
            species=fm.species,
            breed=fm.breed,
            color=fm.color,
            found_location=fm.found_location,
            lat=fm.lat,
            lng=fm.lng,
            photo=fm.photo,
            features=fm.features,
            match_score=fm.match_score,
            created_at=fm.created_at,
            matched_lost=build_lost_response(lr),
        )
        result.append(resp)
    return result


@app.post("/api/lost/{record_id}/reunite", response_model=LostRecordDetail)
def mark_reunited(
    record_id: int,
    found_location: Optional[str] = Form(None),
    photo: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    record = db.query(LostRecord).filter(LostRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    if record.pet.owner_id != current_user.id and not current_user.is_admin:
        raise HTTPException(status_code=403, detail="无权操作")
    record.status = "reunited"
    record.found_time = datetime.utcnow()
    if found_location:
        record.found_location = found_location

    photo_path = save_upload_file(photo) if photo else None

    db.add(ReunionTimeline(
        lost_record_id=record.id,
        event_type="团圆回家",
        description=f"{record.pet.name} 已平安回家！团圆地点：{found_location or '未知'}",
        event_time=datetime.utcnow(),
        photo=photo_path,
    ))
    db.commit()
    db.refresh(record)
    return get_lost_detail(record_id, db)


@app.get("/api/reunited/timeline", response_model=List[LostRecordDetail])
def get_reunited_timeline(db: Session = Depends(get_db)):
    records = db.query(LostRecord).filter(LostRecord.status == "reunited").order_by(LostRecord.found_time.desc()).all()
    result = []
    for r in records:
        resp = build_lost_response(r)
        detail = LostRecordDetail(**resp.model_dump())
        detail.timeline = [TimelineEvent(
            event_type=t.event_type,
            description=t.description,
            event_time=t.event_time,
            photo=t.photo,
        ) for t in sorted(r.timeline, key=lambda x: x.event_time)]
        result.append(detail)
    return result


@app.get("/api/admin/export/pets")
def export_pets_excel(db: Session = Depends(get_db), current_user: User = Depends(get_admin_user)):
    wb = Workbook()
    ws = wb.active
    ws.title = "宠物档案"

    headers = ["ID", "宠物名字", "物种", "品种", "毛色", "芯片号", "是否绝育", "主人联系电话", "主人用户名", "主人邮箱", "描述", "登记时间"]
    ws.append(headers)

    header_fill = PatternFill(start_color="FFF5E6", end_color="FFF5E6", fill_type="solid")
    header_font = Font(bold=True, size=12)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    pets = db.query(Pet).all()
    for pet in pets:
        owner = pet.owner
        ws.append([
            pet.id,
            pet.name,
            pet.species,
            pet.breed,
            pet.color,
            pet.chip_number or "",
            "是" if pet.is_neutered else "否",
            pet.contact_phone,
            owner.username,
            owner.email,
            pet.description or "",
            pet.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=pets_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"},
    )


@app.get("/api/admin/export/lost")
def export_lost_excel(db: Session = Depends(get_db), current_user: User = Depends(get_admin_user)):
    wb = Workbook()
    ws = wb.active
    ws.title = "寻宠记录"

    headers = ["ID", "宠物名字", "物种", "品种", "毛色", "走失时间", "走失地点", "特征描述", "悬赏金额", "状态", "找回时间", "找回地点", "主人", "联系电话"]
    ws.append(headers)

    header_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
    header_font = Font(bold=True, size=12)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    records = db.query(LostRecord).all()
    for r in records:
        pet = r.pet
        owner = pet.owner
        ws.append([
            r.id,
            pet.name,
            pet.species,
            pet.breed,
            pet.color,
            r.lost_time.strftime("%Y-%m-%d %H:%M"),
            r.lost_location,
            r.features,
            r.reward,
            "走失中" if r.status == "lost" else "已回家",
            r.found_time.strftime("%Y-%m-%d %H:%M") if r.found_time else "",
            r.found_location or "",
            owner.username,
            pet.contact_phone,
        ])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=lost_records_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"},
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
